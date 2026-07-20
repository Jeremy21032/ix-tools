#!/usr/bin/env python3
"""
Lista todas las variantes de productos vía Shopify Admin GraphQL (paginado),
incluyendo barcode e inventario por location.

El inventario se pide en una segunda pasada por lotes (nodes) para no superar
el costo máximo por query de Shopify (1000).

Variables de entorno (opcional):
  SHOPIFY_SHOP            ej. tu-store.myshopify.com
  SHOPIFY_ACCESS_TOKEN    token Admin API

Uso:
  py -3 shopify_list_variants.py --shop tu-store.myshopify.com --token shpat_xxx
  py -3 shopify_list_variants.py -o variants.json --excel variants.xlsx
"""

from __future__ import annotations

import argparse
import http.client
import json
import os
import sys
import time
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

# Pass 1: cheap — no inventoryLevels (those explode query cost)
PRODUCTS_QUERY = """
query getProductsWithVariants($cursor: String) {
  products(first: 50, after: $cursor) {
    edges {
      cursor
      node {
        id
        title
        variants(first: 50) {
          pageInfo {
            hasNextPage
            endCursor
          }
          edges {
            node {
              id
              sku
              barcode
              inventoryItem {
                id
              }
            }
          }
        }
      }
    }
    pageInfo {
      hasNextPage
    }
  }
}
"""

VARIANTS_PAGE_QUERY = """
query ProductVariantsPage($id: ID!, $cursor: String!) {
  product(id: $id) {
    id
    title
    variants(first: 50, after: $cursor) {
      pageInfo {
        hasNextPage
        endCursor
      }
      edges {
        node {
          id
          sku
          barcode
          inventoryItem {
            id
          }
        }
      }
    }
  }
}
"""

# Pass 2: inventory by InventoryItem ids (small batches to stay under cost 1000)
INVENTORY_NODES_QUERY = """
query InventoryLevelsByItems($ids: [ID!]!) {
  nodes(ids: $ids) {
    ... on InventoryItem {
      id
      inventoryLevels(first: 25) {
        edges {
          node {
            quantities(names: ["available", "on_hand"]) {
              name
              quantity
            }
            location {
              id
              name
            }
          }
        }
      }
    }
  }
}
"""

# ~8 items × inventoryLevels(first:25) stays safely under max cost 1000
INVENTORY_BATCH_SIZE = 8


def normalize_shop(shop: str) -> str:
    shop = shop.strip()
    if "://" in shop:
        parsed = urlparse(shop)
        host = parsed.netloc or parsed.path
        return host.split("/")[0].strip()
    return shop.rstrip("/")


def post_graphql(
    conn: http.client.HTTPSConnection,
    api_version: str,
    token: str,
    query: str,
    variables: dict[str, Any],
) -> dict[str, Any]:
    path = f"/admin/api/{api_version}/graphql.json"
    body = json.dumps({"query": query, "variables": variables}).encode("utf-8")
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": token,
        "Content-Length": str(len(body)),
        "Accept": "application/json",
    }
    conn.request("POST", path, body=body, headers=headers)
    resp = conn.getresponse()
    raw = resp.read().decode("utf-8")
    if resp.status >= 400:
        raise RuntimeError(f"HTTP {resp.status}: {raw[:800]}")
    return json.loads(raw)


def _ensure_ok(payload: dict[str, Any]) -> dict[str, Any]:
    if payload.get("errors"):
        raise RuntimeError(json.dumps(payload["errors"], indent=2, ensure_ascii=False))
    return payload.get("data") or {}


def _qty_map(level_node: dict[str, Any]) -> dict[str, int | None]:
    out: dict[str, int | None] = {"available": None, "on_hand": None}
    for q in level_node.get("quantities") or []:
        if not isinstance(q, dict):
            continue
        name = str(q.get("name") or "").lower()
        qty = q.get("quantity")
        if name in out:
            try:
                out[name] = int(qty) if qty is not None else None
            except (TypeError, ValueError):
                out[name] = None
    if out["available"] is None and level_node.get("available") is not None:
        try:
            out["available"] = int(level_node["available"])
        except (TypeError, ValueError):
            pass
    return out


def _levels_from_inventory_item(item: dict[str, Any]) -> list[dict[str, Any]]:
    levels_root = item.get("inventoryLevels") or {}
    rows: list[dict[str, Any]] = []
    for edge in levels_root.get("edges") or []:
        node = (edge or {}).get("node") or {}
        loc = node.get("location") or {}
        qtys = _qty_map(node)
        stock = qtys["available"]
        if stock is None:
            stock = qtys["on_hand"]
        rows.append(
            {
                "locationId": loc.get("id") or "",
                "location": loc.get("name") or "",
                "stock": stock if stock is not None else "",
                "onHand": qtys["on_hand"] if qtys["on_hand"] is not None else "",
            }
        )
    return rows


def _collect_variant_nodes(
    conn: http.client.HTTPSConnection,
    api_version: str,
    token: str,
    product: dict[str, Any],
) -> list[dict[str, Any]]:
    block = product.get("variants") or {}
    edges = list(block.get("edges") or [])
    page = block.get("pageInfo") or {}
    cursor = page.get("endCursor")
    while page.get("hasNextPage") and cursor:
        data = _ensure_ok(
            post_graphql(
                conn,
                api_version,
                token,
                VARIANTS_PAGE_QUERY,
                {"id": product.get("id"), "cursor": cursor},
            )
        )
        prod = data.get("product") or {}
        block = prod.get("variants") or {}
        edges.extend(block.get("edges") or [])
        page = block.get("pageInfo") or {}
        cursor = page.get("endCursor")
    return [(e or {}).get("node") or {} for e in edges]


def _fetch_variants_pass(
    conn: http.client.HTTPSConnection,
    api_version: str,
    token: str,
) -> list[dict[str, Any]]:
    """Pass 1: product / variant / sku / barcode / inventoryItemId (no levels)."""
    results: list[dict[str, Any]] = []
    cursor: str | None = None
    has_next = True
    page_n = 0

    while has_next:
        page_n += 1
        data = _ensure_ok(
            post_graphql(conn, api_version, token, PRODUCTS_QUERY, {"cursor": cursor})
        )
        products = data.get("products") or {}
        edges = products.get("edges") or []

        for edge in edges:
            product = edge.get("node") or {}
            for vnode in _collect_variant_nodes(conn, api_version, token, product):
                item = vnode.get("inventoryItem") or {}
                results.append(
                    {
                        "productId": product.get("id"),
                        "productTitle": product.get("title") or "",
                        "variantId": vnode.get("id"),
                        "sku": vnode.get("sku") or "",
                        "barcode": vnode.get("barcode") or "",
                        "inventoryItemId": item.get("id") or "",
                        "inventory": [],
                    }
                )

        page = products.get("pageInfo") or {}
        has_next = bool(page.get("hasNextPage"))
        cursor = edges[-1].get("cursor") if edges else None
        if has_next and not cursor:
            raise RuntimeError("pageInfo.hasNextPage es true pero no hay cursor en el último edge")
        print(f"Productos página {page_n}: variantes acumuladas {len(results)}", flush=True)

    return results


def _fetch_inventory_map(
    conn: http.client.HTTPSConnection,
    api_version: str,
    token: str,
    item_ids: list[str],
) -> dict[str, list[dict[str, Any]]]:
    """Pass 2: map inventoryItemId → levels[]."""
    unique = []
    seen: set[str] = set()
    for iid in item_ids:
        if not iid or iid in seen:
            continue
        seen.add(iid)
        unique.append(iid)

    out: dict[str, list[dict[str, Any]]] = {}
    total = len(unique)
    if not total:
        return out

    for i in range(0, total, INVENTORY_BATCH_SIZE):
        batch = unique[i : i + INVENTORY_BATCH_SIZE]
        data = _ensure_ok(
            post_graphql(
                conn,
                api_version,
                token,
                INVENTORY_NODES_QUERY,
                {"ids": batch},
            )
        )
        for node in data.get("nodes") or []:
            if not node or not isinstance(node, dict):
                continue
            iid = node.get("id")
            if not iid:
                continue
            out[str(iid)] = _levels_from_inventory_item(node)

        done = min(i + INVENTORY_BATCH_SIZE, total)
        print(f"Inventario {done}/{total} items", flush=True)
        # Soft throttle to respect restore rate
        time.sleep(0.15)

    return out


def get_all_product_variants(shop: str, token: str, api_version: str) -> list[dict[str, Any]]:
    """
    Unique variants with nested inventory levels.
    Shape kept compatible with shopify_update_ixc_shipping_metafields
    (productId / variantId / sku still present).
    """
    host = normalize_shop(shop)
    conn = http.client.HTTPSConnection(host, timeout=180)
    try:
        results = _fetch_variants_pass(conn, api_version, token)
        item_ids = [str(r.get("inventoryItemId") or "") for r in results]
        inv_map = _fetch_inventory_map(conn, api_version, token, item_ids)
        for r in results:
            iid = str(r.get("inventoryItemId") or "")
            r["inventory"] = inv_map.get(iid, [])
            # keep payload lean for JSON consumers of nested form
            r.pop("inventoryItemId", None)
        return results
    finally:
        conn.close()


def flatten_variant_rows(variants: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """One row per variant×location (or one empty location row if no inventory)."""
    flat: list[dict[str, Any]] = []
    for v in variants:
        base = {
            "productId": v.get("productId") or "",
            "productTitle": v.get("productTitle") or "",
            "variantId": v.get("variantId") or "",
            "sku": v.get("sku") or "",
            "barcode": v.get("barcode") or "",
        }
        levels = v.get("inventory") or []
        if not levels:
            flat.append(
                {
                    **base,
                    "locationId": "",
                    "location": "",
                    "stock": "",
                    "onHand": "",
                }
            )
            continue
        for lvl in levels:
            flat.append(
                {
                    **base,
                    "locationId": lvl.get("locationId") or "",
                    "location": lvl.get("location") or "",
                    "stock": lvl.get("stock") if lvl.get("stock") != "" else "",
                    "onHand": lvl.get("onHand") if lvl.get("onHand") != "" else "",
                }
            )
    return flat


EXCEL_HEADERS = [
    "productId",
    "productTitle",
    "variantId",
    "sku",
    "barcode",
    "locationId",
    "location",
    "stock",
    "onHand",
]


def write_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise SystemExit(
            "Falta openpyxl. Instale con: py -3 -m pip install openpyxl\n" + str(e)
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Variantes"
    ws.append(EXCEL_HEADERS)
    for r in rows:
        ws.append([r.get(h, "") for h in EXCEL_HEADERS])

    ws.freeze_panes = "A2"
    last_row = max(1, len(rows) + 1)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_HEADERS))}{last_row}"

    widths = (36, 32, 36, 18, 18, 36, 24, 10, 10)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    p = argparse.ArgumentParser(
        description="Shopify Admin GraphQL: variantes + barcode + stock por location."
    )
    p.add_argument(
        "--shop",
        default=os.environ.get("SHOPIFY_SHOP", ""),
        help="Dominio de la tienda (ej. tu-store.myshopify.com)",
    )
    p.add_argument(
        "--token",
        default=os.environ.get("SHOPIFY_ACCESS_TOKEN", ""),
        help="X-Shopify-Access-Token",
    )
    p.add_argument("--api-version", default="2024-10", help="Versión Admin API en la URL")
    p.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Guardar JSON (filas aplanadas: sku, barcode, location, stock)",
    )
    p.add_argument(
        "--excel",
        "-x",
        type=Path,
        help="Guardar Excel (.xlsx) con las mismas columnas",
    )
    p.add_argument(
        "--nested",
        action="store_true",
        help="Guardar JSON anidado (una entrada por variante con inventory[])",
    )
    args = p.parse_args()

    if not args.shop or not args.token:
        print(
            "Faltan credenciales: use --shop y --token o variables "
            "SHOPIFY_SHOP y SHOPIFY_ACCESS_TOKEN.",
            file=sys.stderr,
        )
        return 1

    try:
        data = get_all_product_variants(args.shop, args.token, args.api_version)
    except (OSError, RuntimeError, json.JSONDecodeError) as e:
        print(e, file=sys.stderr)
        return 1

    flat = flatten_variant_rows(data)
    print("TOTAL VARIANTS:", len(data))
    print("TOTAL ROWS (variant×location):", len(flat))
    print(json.dumps(flat[:10], indent=2, ensure_ascii=False))

    if args.output:
        payload = data if args.nested else flat
        args.output.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        print("Guardado JSON:", args.output.resolve())

    if args.excel:
        write_xlsx(args.excel, flat)
        print("Guardado Excel:", args.excel.resolve())

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
