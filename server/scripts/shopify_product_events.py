#!/usr/bin/env python3
"""
shopify_product_events.py

Consulta el objeto Event de la Admin GraphQL API de Shopify para rastrear
cambios de ESTADO (draft/active, publish/unpublish) y ELIMINACION de productos.

LIMITACION CONOCIDA (confirmada por Shopify soporte):
    Este script NO puede rastrear cambios de INVENTARIO/STOCK. Shopify no
    expone via API el historial de ajustes de inventario. La unica fuente
    para eso es el reporte "Inventory adjustment changes" en Analytics >
    Reports del admin (CSV manual).

Uso:
  py -3 shopify_product_events.py --shop mi-tienda.myshopify.com --token shpat_xxx
  py -3 shopify_product_events.py --shop mi-tienda --token shpat_xxx --actions destroy
  py -3 shopify_product_events.py --shop mi-tienda --token shpat_xxx -o eventos.csv --excel eventos.xlsx

Requiere Admin API token con scope read_products (y acceso a events).
"""

from __future__ import annotations

import argparse
import csv
import http.client
import json
import os
import sys
import time
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

EVENTS_QUERY = """
query GetProductEvents($query: String!, $cursor: String) {
  events(query: $query, first: 100, after: $cursor, sortKey: CREATED_AT) {
    edges {
      cursor
      node {
        id
        createdAt
        message
        ... on BasicEvent {
          action
          subjectType
          appTitle
          attributeToApp
          criticalAlert
          subject {
            __typename
            ... on Product {
              id
              title
              handle
              status
            }
          }
        }
        ... on CommentEvent {
          rawMessage
        }
      }
    }
    pageInfo {
      hasNextPage
    }
  }
}
"""

CSV_HEADERS = [
    "createdAt",
    "action",
    "message",
    "productId",
    "productTitle",
    "productHandle",
    "productStatus",
    "appTitle",
    "criticalAlert",
    "eventId",
]


def normalize_shop(shop: str) -> str:
    shop = (shop or "").strip()
    if "://" in shop:
        parsed = urlparse(shop)
        host = parsed.netloc or parsed.path
        return host.split("/")[0].strip()
    shop = shop.rstrip("/")
    if not shop.endswith(".myshopify.com") and "." not in shop:
        return f"{shop}.myshopify.com"
    return shop


def post_graphql(
    conn: http.client.HTTPSConnection,
    api_version: str,
    token: str,
    query: str,
    variables: dict[str, Any],
) -> dict[str, Any]:
    path = f"/admin/api/{api_version}/graphql.json"
    body = json.dumps({"query": query, "variables": variables}).encode("utf-8")
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": token,
        "Content-Length": str(len(body)),
        "Accept": "application/json",
    }
    conn.request("POST", path, body=body, headers=headers)
    resp = conn.getresponse()
    raw = resp.read().decode("utf-8")
    if resp.status >= 400:
        raise RuntimeError(f"HTTP {resp.status}: {raw[:800]}")
    return json.loads(raw)


def flatten_event(node: dict[str, Any]) -> dict[str, Any]:
    subject = node.get("subject") or {}
    return {
        "createdAt": node.get("createdAt") or "",
        "action": node.get("action") or "",
        "message": node.get("message") or node.get("rawMessage") or "",
        "productId": subject.get("id") or "",
        "productTitle": subject.get("title") or "",
        "productHandle": subject.get("handle") or "",
        "productStatus": subject.get("status") or "",
        "appTitle": node.get("appTitle") or "",
        "criticalAlert": node.get("criticalAlert") if node.get("criticalAlert") is not None else "",
        "eventId": node.get("id") or "",
    }


def fetch_product_events(
    shop: str,
    token: str,
    api_version: str,
    actions: list[str] | None = None,
) -> list[dict[str, Any]]:
    """
    actions: create | update | destroy | publish | unpublish (según lo que Shopify exponga).
    Filtro base: subject_type PRODUCT.
    """
    filters = ["subject_type:'PRODUCT'"]
    if actions:
        cleaned = [a.strip() for a in actions if a and str(a).strip()]
        if cleaned:
            action_filter = " OR ".join([f"action:'{a}'" for a in cleaned])
            filters.append(f"({action_filter})")

    query_string = " AND ".join(filters)
    host = normalize_shop(shop)
    cursor: str | None = None
    rows: list[dict[str, Any]] = []
    page = 0

    conn = http.client.HTTPSConnection(host, timeout=120)
    try:
        while True:
            page += 1
            payload = post_graphql(
                conn,
                api_version,
                token,
                EVENTS_QUERY,
                {"query": query_string, "cursor": cursor},
            )
            if payload.get("errors"):
                raise RuntimeError(json.dumps(payload["errors"], indent=2, ensure_ascii=False))

            data = (payload.get("data") or {}).get("events") or {}
            edges = data.get("edges") or []
            for edge in edges:
                node = (edge or {}).get("node") or {}
                rows.append(flatten_event(node))

            print(f"Página {page}: {len(rows)} eventos acumulados", flush=True)

            page_info = data.get("pageInfo") or {}
            if not page_info.get("hasNextPage") or not edges:
                break
            cursor = edges[-1].get("cursor")
            if not cursor:
                break
            time.sleep(0.5)
    finally:
        conn.close()

    return rows


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS, extrasaction="ignore")
        writer.writeheader()
        for r in rows:
            writer.writerow({h: r.get(h, "") for h in CSV_HEADERS})


def write_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise SystemExit(
            "Falta openpyxl. Instale con: py -3 -m pip install openpyxl\n" + str(e)
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Product events"
    ws.append(CSV_HEADERS)
    for r in rows:
        ws.append([r.get(h, "") for h in CSV_HEADERS])
    ws.freeze_panes = "A2"
    last = max(1, len(rows) + 1)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CSV_HEADERS))}{last}"
    widths = (22, 12, 48, 40, 28, 22, 12, 18, 12, 40)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    p = argparse.ArgumentParser(
        description="Rastrea eventos de productos Shopify (estado/eliminación). No cubre inventario."
    )
    p.add_argument(
        "--shop",
        default=os.environ.get("SHOPIFY_SHOP", ""),
        help="Tienda: mi-tienda o mi-tienda.myshopify.com",
    )
    p.add_argument(
        "--token",
        default=os.environ.get("SHOPIFY_ACCESS_TOKEN", ""),
        help="X-Shopify-Access-Token",
    )
    p.add_argument("--api-version", default="2025-01", help="Versión Admin API")
    p.add_argument(
        "--actions",
        nargs="*",
        default=None,
        help="Filtrar acciones, ej: create update destroy publish unpublish",
    )
    p.add_argument("-o", "--output", type=Path, default=Path("product_events.csv"), help="CSV salida")
    p.add_argument("--excel", "-x", type=Path, help="Excel (.xlsx) opcional")
    args = p.parse_args()

    if not args.shop or not args.token:
        print("Faltan --shop y --token (o SHOPIFY_SHOP / SHOPIFY_ACCESS_TOKEN).", file=sys.stderr)
        return 1

    try:
        rows = fetch_product_events(args.shop, args.token, args.api_version, actions=args.actions)
    except (OSError, RuntimeError, json.JSONDecodeError) as e:
        print(e, file=sys.stderr)
        return 1

    print(f"TOTAL EVENTOS PRODUCT: {len(rows)}", flush=True)
    write_csv(args.output, rows)
    print(f"Guardado CSV: {args.output.resolve()}", flush=True)
    if args.excel:
        write_xlsx(args.excel, rows)
        print(f"Guardado Excel: {args.excel.resolve()}", flush=True)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
