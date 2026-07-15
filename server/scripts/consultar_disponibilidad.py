#!/usr/bin/env python3
"""
Consulta disponibilidad de SKUs en la API IXC y exporta resultados a Excel.

Requiere: pip install openpyxl requests

Uso:
  py -3 consultar_disponibilidad.py --help
  py -3 consultar_disponibilidad.py --skus skus.txt
  py -3 consultar_disponibilidad.py --skus skus.txt -o disponibilidad.xlsx
  py -3 consultar_disponibilidad.py --skus skusKCCR.txt -o disponibilidad_kccr.xlsx
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
import time

import requests

DEFAULT_URL = (
    "https://availability.eastus2.cloudapp.azure.com/logic/products/availabilities"
)
DEFAULT_HEADERS = {
    "nameCode": "skuIxc",
    "x-api-version": "1",
    "offset": "0",
    "limit": "1000",
}
# La API devuelve 404 si se envian mas de ~20 SKUs en una sola URL.
MAX_SKUS_PER_REQUEST = 20


def load_skus_from_file(path: Path) -> tuple[list[str], list[str], dict[str, list[int]]]:
    """
    Lee el archivo y devuelve:
      - all_skus: todos los SKU en orden (incluye duplicados)
      - unique_skus: unicos para consultar la API (orden de primera aparicion)
      - sku_lines: mapa sku -> numeros de fila en el archivo (1-based)
    """
    all_skus: list[str] = []
    sku_lines: dict[str, list[int]] = {}
    fila = 0
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        for part in line.replace(",", " ").split():
            sku = part.strip()
            if not sku:
                continue
            fila += 1
            all_skus.append(sku)
            sku_lines.setdefault(sku, []).append(fila)
    unique_skus = list(dict.fromkeys(all_skus))
    return all_skus, unique_skus, sku_lines


def load_skus(path: Path) -> list[str]:
    """Solo SKUs unicos (compatibilidad)."""
    _, unique, _ = load_skus_from_file(path)
    return unique


def chunk_list(items: list[str], size: int) -> list[list[str]]:
    return [items[i : i + size] for i in range(0, len(items), size)]


def merge_payloads(left: dict[str, Any], right: dict[str, Any]) -> dict[str, Any]:
    return {
        "transactionDateTime": right.get("transactionDateTime")
        or left.get("transactionDateTime"),
        "products": (left.get("products") or []) + (right.get("products") or []),
        "skuNotFound": (left.get("skuNotFound") or []) + (right.get("skuNotFound") or []),
    }


def fetch_batch(
    url: str,
    headers: dict[str, str],
    skus: list[str],
    timeout: int,
) -> dict[str, Any]:
    params = [("sku", sku) for sku in skus]
    response = requests.get(url, headers=headers, params=params, timeout=timeout)
    response.raise_for_status()
    return response.json()


def fetch_skus_safe(
    url: str,
    headers: dict[str, str],
    skus: list[str],
    timeout: int,
) -> dict[str, Any]:
    """Consulta SKUs; si el lote es demasiado grande (404), divide y reintenta."""
    if not skus:
        return {"products": [], "skuNotFound": []}
    try:
        return fetch_batch(url, headers, skus, timeout)
    except requests.HTTPError as exc:
        status = exc.response.status_code if exc.response is not None else None
        if len(skus) == 1 or status not in (404, 413, 414):
            raise
        mid = len(skus) // 2
        left = fetch_skus_safe(url, headers, skus[:mid], timeout)
        right = fetch_skus_safe(url, headers, skus[mid:], timeout)
        return merge_payloads(left, right)


def common_code(product: dict[str, Any], name: str) -> str:
    for item in product.get("commonCodes") or []:
        if item.get("nameCode") == name:
            return str(item.get("codeValue") or "")
    return ""


def parse_products(payload: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str]]:
    products = payload.get("products") or []
    not_found = [str(s) for s in (payload.get("skuNotFound") or [])]
    return products, not_found


def flatten_product_rows(product: dict[str, Any]) -> tuple[list[dict], list[dict]]:
    sku = (product.get("codeProduct") or {}).get("codeValue", "")
    status = product.get("status", "")
    mpn = common_code(product, "MPN")
    upc = common_code(product, "UPC")
    ixc = common_code(product, "IXC")

    country_rows: list[dict[str, Any]] = []
    location_rows: list[dict[str, Any]] = []

    for country in product.get("countries") or []:
        country_id = country.get("countryId", "")
        qty = country.get("quantityinStock") or {}
        country_rows.append(
            {
                "sku": sku,
                "status": status,
                "mpn": mpn,
                "upc": upc,
                "ixc": ixc,
                "countryId": country_id,
                "OH": qty.get("OH", 0),
                "OS": qty.get("OS", 0),
                "AFS": qty.get("AFS", 0),
                "ORW": qty.get("ORW", 0),
                "OPO": qty.get("OPO", 0),
            }
        )
        for loc in country.get("inStock") or []:
            location_rows.append(
                {
                    "sku": sku,
                    "status": status,
                    "countryId": country_id,
                    "locationId": loc.get("locationId", ""),
                    "locationName": loc.get("locationName", ""),
                    "locationAddress": loc.get("locationAddress", ""),
                    "OH": loc.get("OH", 0),
                    "OS": loc.get("OS", 0),
                    "AFS": loc.get("AFS", 0),
                    "ORW": loc.get("ORW", 0),
                    "OPO": loc.get("OPO", 0),
                }
            )

    return country_rows, location_rows


def lookup_sku_result(
    sku: str,
    found_by_sku: dict[str, dict[str, Any]],
    not_found_set: set[str],
) -> dict[str, str]:
    if sku in not_found_set:
        return {
            "encontrado": "No",
            "status": "",
            "mpn": "",
            "upc": "",
            "ixc": "",
        }
    if sku in found_by_sku:
        product = found_by_sku[sku]
        return {
            "encontrado": "Si",
            "status": str(product.get("status", "")),
            "mpn": common_code(product, "MPN"),
            "upc": common_code(product, "UPC"),
            "ixc": common_code(product, "IXC"),
        }
    return {
        "encontrado": "No",
        "status": "",
        "mpn": "",
        "upc": "",
        "ixc": "",
    }


def build_summary_rows(
    all_skus: list[str],
    sku_lines: dict[str, list[int]],
    products: list[dict[str, Any]],
    not_found: list[str],
) -> list[dict[str, Any]]:
    found_by_sku = {
        (p.get("codeProduct") or {}).get("codeValue", ""): p for p in products
    }
    not_found_set = set(not_found)
    rows: list[dict[str, Any]] = []

    for fila, sku in enumerate(all_skus, start=1):
        veces = len(sku_lines.get(sku, [fila]))
        result = lookup_sku_result(sku, found_by_sku, not_found_set)
        rows.append(
            {
                "fila": fila,
                "sku": sku,
                "veces_en_archivo": veces,
                "duplicado_en_archivo": "Si" if veces > 1 else "No",
                **result,
            }
        )
    return rows


def build_duplicate_rows(sku_lines: dict[str, list[int]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sku, filas in sorted(sku_lines.items()):
        if len(filas) > 1:
            rows.append(
                {
                    "sku": sku,
                    "veces": len(filas),
                    "filas": ", ".join(str(f) for f in filas),
                }
            )
    return rows


def write_sheet(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def export_excel(
    output_path: Path,
    summary_rows: list[dict[str, Any]],
    not_found_rows: list[dict[str, Any]],
    duplicate_rows: list[dict[str, Any]],
    country_rows: list[dict[str, Any]],
    location_rows: list[dict[str, Any]],
    meta: dict[str, str],
) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError(
            "Falta openpyxl. Instale con: py -3 -m pip install openpyxl"
        ) from exc

    wb = Workbook()

    ws_meta = wb.active
    ws_meta.title = "Info"
    for key, value in meta.items():
        ws_meta.append([key, value])

    ws_resumen = wb.create_sheet("Resumen")
    write_sheet(
        ws_resumen,
        [
            "fila",
            "sku",
            "veces_en_archivo",
            "duplicado_en_archivo",
            "encontrado",
            "status",
            "mpn",
            "upc",
            "ixc",
        ],
        summary_rows,
    )

    ws_nf = wb.create_sheet("No encontrados")
    write_sheet(ws_nf, ["sku"], [{"sku": s} for s in not_found_rows])

    ws_dup = wb.create_sheet("Duplicados en archivo")
    write_sheet(ws_dup, ["sku", "veces", "filas"], duplicate_rows)

    ws_country = wb.create_sheet("Stock por pais")
    write_sheet(
        ws_country,
        ["sku", "status", "mpn", "upc", "ixc", "countryId", "OH", "OS", "AFS", "ORW", "OPO"],
        country_rows,
    )

    ws_loc = wb.create_sheet("Stock por ubicacion")
    write_sheet(
        ws_loc,
        [
            "sku",
            "status",
            "countryId",
            "locationId",
            "locationName",
            "locationAddress",
            "OH",
            "OS",
            "AFS",
            "ORW",
            "OPO",
        ],
        location_rows,
    )

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    epilog = """
Ejemplos:
  py -3 consultar_disponibilidad.py --help
  py -3 consultar_disponibilidad.py
  py -3 consultar_disponibilidad.py --skus skus.txt
  py -3 consultar_disponibilidad.py --skus mi_lista.txt -o reporte.xlsx
  py -3 consultar_disponibilidad.py --skus skusKCCR.txt -o disponibilidad_kccr.xlsx

Limite de la API:
  Maximo ~20 SKUs por request. Con --batch-size mayor, el script divide
  automaticamente si recibe error 404.

Archivo de SKUs (skus.txt):
  Un SKU por linea. Tambien varios en la misma linea separados por coma o espacio.
  Las lineas que empiezan con # se ignoran.

Excel generado (hojas):
  Info                 Metadatos (lineas, unicos, duplicados)
  Resumen              Una fila por cada linea del archivo (incluye duplicados)
  No encontrados       SKUs que la API devolvio en skuNotFound
  Duplicados en archivo SKUs repetidos en tu lista y en que filas
  Stock por pais       Inventario agregado por pais
  Stock por ubicacion  Detalle por tienda/bodega

Nota:
  Si el archivo tiene SKUs repetidos, la API se consulta solo una vez por SKU
  unico, pero el Resumen repite el resultado en cada fila del archivo.

Dependencias:
  py -3 -m pip install openpyxl requests
"""
    parser = argparse.ArgumentParser(
        prog="consultar_disponibilidad.py",
        description=(
            "Consulta disponibilidad de SKUs en la API IXC y exporta un Excel "
            "para comparar stock y detectar SKUs no encontrados."
        ),
        epilog=epilog,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--skus",
        "--input",
        dest="skus",
        default="skus.txt",
        metavar="ARCHIVO",
        help="Archivo con SKUs a consultar (default: skus.txt).",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="",
        metavar="XLSX",
        help=(
            "Ruta del Excel de salida. "
            "Si no se indica, se crea disponibilidad_YYYYMMDD_HHMMSS.xlsx."
        ),
    )
    parser.add_argument(
        "--url",
        default=DEFAULT_URL,
        metavar="URL",
        help="URL del endpoint de disponibilidad.",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=MAX_SKUS_PER_REQUEST,
        metavar="N",
        help=f"SKUs por lote (max recomendado: {MAX_SKUS_PER_REQUEST}, default: {MAX_SKUS_PER_REQUEST}).",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=60,
        metavar="SEG",
        help="Timeout de cada request en segundos (default: 60).",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=0.0,
        metavar="SEG",
        help="Pausa entre lotes en segundos (default: 0).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    skus_path = Path(args.skus)
    if not skus_path.is_file():
        print(f"No existe el archivo de SKUs: {skus_path}", file=sys.stderr)
        return 1

    all_skus, skus, sku_lines = load_skus_from_file(skus_path)
    if not skus:
        print(f"No hay SKUs en {skus_path}", file=sys.stderr)
        return 1

    duplicados_extra = len(all_skus) - len(skus)
    if duplicados_extra > 0:
        print(
            f"Archivo: {len(all_skus)} linea(s), {len(skus)} SKU(s) unicos, "
            f"{duplicados_extra} linea(s) duplicada(s) (no se consultan de nuevo)."
        )

    if args.output:
        output_path = Path(args.output)
    else:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path(__file__).resolve().parent / f"disponibilidad_{stamp}.xlsx"

    all_products: list[dict[str, Any]] = []
    all_not_found: list[str] = []
    last_transaction = ""

    batch_size = max(1, args.batch_size)
    if batch_size > MAX_SKUS_PER_REQUEST:
        print(
            f"Aviso: --batch-size {batch_size} supera el limite de la API "
            f"({MAX_SKUS_PER_REQUEST}). Se usara {MAX_SKUS_PER_REQUEST}.",
            file=sys.stderr,
        )
        batch_size = MAX_SKUS_PER_REQUEST

    batches = chunk_list(skus, batch_size)
    print(f"Consultando {len(skus)} SKU(s) unicos en {len(batches)} lote(s)...")

    for index, batch in enumerate(batches, start=1):
        print(f"  Lote {index}/{len(batches)} ({len(batch)} SKU(s))...", flush=True)
        payload = fetch_skus_safe(args.url, DEFAULT_HEADERS, batch, args.timeout)
        if args.delay > 0 and index < len(batches):
            time.sleep(args.delay)
        products, not_found = parse_products(payload)
        all_products.extend(products)
        all_not_found.extend(not_found)
        last_transaction = str(payload.get("transactionDateTime") or last_transaction)

    country_rows: list[dict[str, Any]] = []
    location_rows: list[dict[str, Any]] = []
    for product in all_products:
        c_rows, l_rows = flatten_product_rows(product)
        country_rows.extend(c_rows)
        location_rows.extend(l_rows)

    not_found_unique = sorted(set(all_not_found))
    summary_rows = build_summary_rows(all_skus, sku_lines, all_products, not_found_unique)
    duplicate_rows = build_duplicate_rows(sku_lines)
    found_by_sku = {
        (p.get("codeProduct") or {}).get("codeValue", ""): p for p in all_products
    }
    not_found_set = set(not_found_unique)
    unicos_encontrados = sum(
        1
        for sku in skus
        if lookup_sku_result(sku, found_by_sku, not_found_set)["encontrado"] == "Si"
    )

    meta = {
        "archivo_skus": str(skus_path.resolve()),
        "total_lineas_archivo": str(len(all_skus)),
        "skus_unicos": str(len(skus)),
        "lineas_duplicadas_en_archivo": str(duplicados_extra),
        "skus_unicos_encontrados": str(unicos_encontrados),
        "skus_unicos_no_encontrados": str(len(not_found_unique)),
        "filas_resumen": str(len(summary_rows)),
        "transactionDateTime": last_transaction,
        "generado_en": datetime.now(timezone.utc).isoformat(),
        "url": args.url,
        "batch_size": str(batch_size),
    }

    export_excel(
        output_path,
        summary_rows,
        not_found_unique,
        duplicate_rows,
        country_rows,
        location_rows,
        meta,
    )

    print(f"\nListo: {output_path}")
    print(f"  Filas en Resumen: {meta['filas_resumen']}")
    print(f"  SKUs unicos encontrados: {meta['skus_unicos_encontrados']}")
    print(f"  SKUs unicos no encontrados (API): {meta['skus_unicos_no_encontrados']}")
    if duplicate_rows:
        print(f"  SKUs duplicados en archivo: {len(duplicate_rows)}")
    if not_found_unique:
        print("  SKUs no encontrados:", ", ".join(not_found_unique))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
