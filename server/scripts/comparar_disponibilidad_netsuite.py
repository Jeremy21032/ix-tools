#!/usr/bin/env python3
"""
Compara disponibilidad API (Excel) vs export NetSuite (CSV) por ubicacion o por pais.

Requiere: pip install openpyxl

Uso:
  py -3 comparar_disponibilidad_netsuite.py --help
  py -3 comparar_disponibilidad_netsuite.py \\
      --api disponibilidad_kccr_v2.xlsx \\
      --netsuite "C:/Users/Usuario/Downloads/availabilityNetsuite_CR_20260515_JLE.csv" \\
      --country CR \\
      -o comparacion_cr.xlsx
"""

from __future__ import annotations

import argparse
import csv
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

QTY_FIELDS = ("OH", "OS", "AFS", "ORW", "OPO")
NS_TO_API = {
    "ohQty": "OH",
    "osoQty": "OS",
    "afsQty": "AFS",
    "rwQty": "ORW",
    "opoQty": "OPO",
}


def to_int(value: Any) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def load_api_ubicacion(path: Path, country: str) -> dict[tuple[str, str, str], dict[str, Any]]:
    from openpyxl import load_workbook

    rows: dict[tuple[str, str, str], dict[str, Any]] = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Stock por ubicacion" not in wb.sheetnames:
        raise ValueError(f"El Excel no tiene hoja 'Stock por ubicacion': {path}")
    ws = wb["Stock por ubicacion"]
    header = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(max_row=1, values_only=True))]
    idx = {name: header.index(name) for name in header}

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[idx["sku"]] or "").strip()
        country_id = str(row[idx["countryId"]] or "").strip()
        location_id = str(row[idx["locationId"]] or "").strip()
        if not sku or (country and country_id != country):
            continue
        key = (sku, country_id, location_id)
        rows[key] = {
            "sku": sku,
            "countryId": country_id,
            "locationId": location_id,
            "locationName": row[idx.get("locationName", idx["locationId"])],
            "OH": to_int(row[idx["OH"]]),
            "OS": to_int(row[idx["OS"]]),
            "AFS": to_int(row[idx["AFS"]]),
            "ORW": to_int(row[idx["ORW"]]),
            "OPO": to_int(row[idx["OPO"]]),
        }
    wb.close()
    return rows


def load_api_pais(path: Path, country: str) -> dict[tuple[str, str], dict[str, Any]]:
    from openpyxl import load_workbook

    rows: dict[tuple[str, str], dict[str, Any]] = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Stock por pais" not in wb.sheetnames:
        raise ValueError(f"El Excel no tiene hoja 'Stock por pais': {path}")
    ws = wb["Stock por pais"]
    header = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(max_row=1, values_only=True))]
    idx = {name: header.index(name) for name in header}

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[idx["sku"]] or "").strip()
        country_id = str(row[idx["countryId"]] or "").strip()
        if not sku or (country and country_id != country):
            continue
        key = (sku, country_id)
        rows[key] = {
            "sku": sku,
            "countryId": country_id,
            **{f: to_int(row[idx[f]]) for f in QTY_FIELDS},
        }
    wb.close()
    return rows


def load_netsuite_ubicacion(path: Path, country: str) -> dict[tuple[str, str, str], dict[str, Any]]:
    rows: dict[tuple[str, str, str], dict[str, Any]] = {}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            sku = str(row.get("skuIx") or row.get("sku") or "").strip()
            country_id = str(row.get("country") or "").strip()
            location_id = str(row.get("locationId") or "").strip()
            if not sku or (country and country_id != country):
                continue
            key = (sku, country_id, location_id)
            rows[key] = {
                "sku": sku,
                "countryId": country_id,
                "locationId": location_id,
                "location": row.get("location", ""),
                "OH": to_int(row.get("ohQty")),
                "OS": to_int(row.get("osoQty")),
                "AFS": to_int(row.get("afsQty")),
                "ORW": to_int(row.get("rwQty")),
                "OPO": to_int(row.get("opoQty")),
            }
    return rows


def aggregate_netsuite_pais(
    ubicacion_rows: dict[tuple[str, str, str], dict[str, Any]],
) -> dict[tuple[str, str], dict[str, Any]]:
    aggregated: dict[tuple[str, str], dict[str, Any]] = {}
    for (sku, country_id, _location_id), row in ubicacion_rows.items():
        key = (sku, country_id)
        if key not in aggregated:
            aggregated[key] = {
                "sku": sku,
                "countryId": country_id,
                **{f: 0 for f in QTY_FIELDS},
            }
        for field in QTY_FIELDS:
            aggregated[key][field] += row[field]
    return aggregated


def qty_diff(api_val: int, ns_val: int) -> int:
    return api_val - ns_val


def compare_rows(
    api_rows: dict[tuple, dict[str, Any]],
    ns_rows: dict[tuple, dict[str, Any]],
    key_fields: list[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    all_keys = sorted(set(api_rows) | set(ns_rows))
    diferencias: list[dict[str, Any]] = []
    coinciden: list[dict[str, Any]] = []
    solo_api: list[dict[str, Any]] = []
    solo_ns: list[dict[str, Any]] = []

    for key in all_keys:
        api = api_rows.get(key)
        ns = ns_rows.get(key)
        base: dict[str, Any] = {}
        for field in key_fields:
            if api and field in api:
                base[field] = api[field]
            elif ns and field in ns:
                base[field] = ns[field]
        if api:
            base.setdefault("locationName", api.get("locationName", ""))
        if ns:
            base.setdefault("location", ns.get("location", ""))

        if api and not ns:
            solo_row = {**base, "estado": "Solo en API"}
            for field in QTY_FIELDS:
                solo_row[f"api_{field}"] = api[field]
            solo_api.append(solo_row)
            continue
        if ns and not api:
            solo_row = {**base, "estado": "Solo en NetSuite"}
            for field in QTY_FIELDS:
                solo_row[f"ns_{field}"] = ns[field]
            solo_ns.append(solo_row)
            continue

        row: dict[str, Any] = {**base, "estado": "OK"}
        has_diff = False
        for field in QTY_FIELDS:
            api_val = api[field]
            ns_val = ns[field]
            row[f"api_{field}"] = api_val
            row[f"ns_{field}"] = ns_val
            diff = qty_diff(api_val, ns_val)
            row[f"diff_{field}"] = diff
            if diff != 0:
                has_diff = True

        if has_diff:
            row["estado"] = "Diferencia"
            diferencias.append(row)
        else:
            coinciden.append(row)

    return diferencias, coinciden, solo_api, solo_ns


def write_sheet(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def export_excel(
    output_path: Path,
    nivel: str,
    diferencias: list[dict[str, Any]],
    coinciden: list[dict[str, Any]],
    solo_api: list[dict[str, Any]],
    solo_ns: list[dict[str, Any]],
    meta: dict[str, str],
) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError(
            "Falta openpyxl. Instale con: py -3 -m pip install openpyxl"
        ) from exc

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "Info"
    for key, value in meta.items():
        ws_info.append([key, value])

    if nivel == "ubicacion":
        key_headers = ["sku", "countryId", "locationId"]
        extra = ["locationName", "location"]
    else:
        key_headers = ["sku", "countryId"]
        extra = []

    diff_headers = key_headers + extra + [
        "estado",
        *[x for f in QTY_FIELDS for x in (f"api_{f}", f"ns_{f}", f"diff_{f}")],
    ]
    simple_headers = key_headers + extra + ["estado"]

    ws_diff = wb.create_sheet("Diferencias")
    write_sheet(ws_diff, diff_headers, diferencias)

    ws_ok = wb.create_sheet("Coinciden")
    write_sheet(ws_ok, diff_headers, coinciden)

    solo_api_headers = simple_headers + [f"api_{f}" for f in QTY_FIELDS]
    solo_ns_headers = simple_headers + [f"ns_{f}" for f in QTY_FIELDS]

    ws_api = wb.create_sheet("Solo en API")
    write_sheet(ws_api, solo_api_headers, solo_api)

    ws_ns = wb.create_sheet("Solo en NetSuite")
    write_sheet(ws_ns, solo_ns_headers, solo_ns)

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    epilog = """
Ejemplos:
  py -3 comparar_disponibilidad_netsuite.py --help
  py -3 comparar_disponibilidad_netsuite.py \\
      --api disponibilidad_kccr_v2.xlsx \\
      --netsuite ../availabilityNetsuite_CR_20260515_JLE.csv \\
      --country CR -o comparacion_cr.xlsx

  py -3 comparar_disponibilidad_netsuite.py \\
      --api disponibilidad_kccr_v2.xlsx \\
      --netsuite ../availabilityNetsuite_CR_20260515_JLE.csv \\
      --nivel pais --country CR

Columnas comparadas (API vs NetSuite):
  OH  <-> ohQty
  OS  <-> osoQty
  AFS <-> afsQty
  ORW <-> rwQty
  OPO <-> opoQty

diff_* = valor API - valor NetSuite (positivo = API tiene mas)
"""
    parser = argparse.ArgumentParser(
        prog="comparar_disponibilidad_netsuite.py",
        description="Compara stock API vs CSV NetSuite y exporta diferencias a Excel.",
        epilog=epilog,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--api",
        "--input",
        dest="api",
        required=True,
        metavar="XLSX",
        help="Excel generado por consultar_disponibilidad.py",
    )
    parser.add_argument(
        "--netsuite",
        required=True,
        metavar="CSV",
        help="CSV de disponibilidad NetSuite.",
    )
    parser.add_argument(
        "--country",
        default="CR",
        metavar="PAIS",
        help="Filtrar por codigo de pais (default: CR). Use vacio para todos.",
    )
    parser.add_argument(
        "--nivel",
        choices=("ubicacion", "pais"),
        default="ubicacion",
        help="Comparar por ubicacion (default) o agregado por pais.",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="",
        metavar="XLSX",
        help="Excel de salida (default: comparacion_PAIS_TIMESTAMP.xlsx).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    api_path = Path(args.api)
    ns_path = Path(args.netsuite)
    country = args.country.strip().upper() if args.country else ""

    if not api_path.is_file():
        print(f"No existe el Excel API: {api_path}", file=sys.stderr)
        return 1
    if not ns_path.is_file():
        print(f"No existe el CSV NetSuite: {ns_path}", file=sys.stderr)
        return 1

    if args.output:
        output_path = Path(args.output)
    else:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = country or "ALL"
        output_path = Path(__file__).resolve().parent / f"comparacion_{suffix}_{stamp}.xlsx"

    print(f"Comparando nivel={args.nivel}, pais={country or 'TODOS'}...")

    if args.nivel == "ubicacion":
        api_rows = load_api_ubicacion(api_path, country)
        ns_rows = load_netsuite_ubicacion(ns_path, country)
        key_fields = ["sku", "countryId", "locationId"]
        for key, row in api_rows.items():
            row["locationName"] = row.get("locationName", "")
        for key, row in ns_rows.items():
            row["location"] = row.get("location", "")
    else:
        api_rows = load_api_pais(api_path, country)
        ns_rows = aggregate_netsuite_pais(load_netsuite_ubicacion(ns_path, country))
        key_fields = ["sku", "countryId"]

    diferencias, coinciden, solo_api, solo_ns = compare_rows(api_rows, ns_rows, key_fields)

    meta = {
        "excel_api": str(api_path.resolve()),
        "csv_netsuite": str(ns_path.resolve()),
        "nivel": args.nivel,
        "pais_filtro": country or "TODOS",
        "registros_api": str(len(api_rows)),
        "registros_netsuite": str(len(ns_rows)),
        "coinciden": str(len(coinciden)),
        "con_diferencias": str(len(diferencias)),
        "solo_en_api": str(len(solo_api)),
        "solo_en_netsuite": str(len(solo_ns)),
        "generado_en": datetime.now(timezone.utc).isoformat(),
    }

    export_excel(output_path, args.nivel, diferencias, coinciden, solo_api, solo_ns, meta)

    print(f"\nListo: {output_path}")
    print(f"  Coinciden: {meta['coinciden']}")
    print(f"  Con diferencias de cantidad: {meta['con_diferencias']}")
    print(f"  Solo en API: {meta['solo_en_api']}")
    print(f"  Solo en NetSuite: {meta['solo_en_netsuite']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
