#!/usr/bin/env python3
"""
Exporta todos los productos y sus variantes desde Shopify Admin GraphQL a un Excel (.xlsx)
para el área comercial, incluyendo metafields del namespace IXC (shipping y el resto en JSON).

Requiere: pip install openpyxl

Variables de entorno: SHOPIFY_SHOP, SHOPIFY_ACCESS_TOKEN (o --shop / --token).

Uso:
  py -3 shopify_products_variants_excel.py --shop tu-tienda.myshopify.com --token shpat_xxx
  py -3 shopify_products_variants_excel.py -o catalogo.xlsx
  py -3 shopify_products_variants_excel.py --help
"""

from __future__ import annotations

import argparse
import http.client
import json
import os
import re
import sys
import time
from pathlib import Path
from typing import Any

_ROOT = str(Path(__file__).resolve().parent)
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

from shopify_list_variants import post_graphql, normalize_shop

# Mismo orden que shopify_update_ixc_shipping_metafields.py (namespace IXC)
IXC_SHIPPING_KEYS = [
    "shipping_LengthEach",
    "shipping_widthEach",
    "shipping_heightEach",
    "shipping_weightEach",
    "shipping_volumeEach",
]

METAFIELDS_FRAGMENT = """
              metafields(first: 40, namespace: "IXC") {
                edges {
                  node {
                    key
                    value
                    type
                  }
                }
              }
"""

PRODUCTS_PAGE_QUERY = """
query ProductsForExport($cursor: String) {
  products(first: 50, after: $cursor) {
    pageInfo {
      hasNextPage
    }
    edges {
      cursor
      node {
        id
        title
        handle
        status
        vendor
        productType
        variants(first: 250) {
          pageInfo {
            hasNextPage
            endCursor
          }
          edges {
            node {
              id
              title
              sku
              barcode
              price
              compareAtPrice
              inventoryQuantity
              selectedOptions {
                name
                value
              }
""" + METAFIELDS_FRAGMENT + """
            }
          }
        }
      }
    }
  }
}
"""

VARIANTS_PAGE_QUERY = """
query ProductVariantsPage($id: ID!, $cursor: String!) {
  product(id: $id) {
    id
    variants(first: 250, after: $cursor) {
      pageInfo {
        hasNextPage
        endCursor
      }
      edges {
        node {
          id
          title
          sku
          barcode
          price
          compareAtPrice
          inventoryQuantity
          selectedOptions {
            name
            value
          }
""" + METAFIELDS_FRAGMENT + """
        }
      }
    }
  }
}
"""

HEADERS_ES = [
    "ID producto",
    "Producto",
    "Handle",
    "Estado",
    "Vendor",
    "Tipo producto",
    "ID variante",
    "Título variante",
    "SKU",
    "Código de barras",
    "Opción 1",
    "Opción 2",
    "Opción 3",
    "Precio",
    "Precio comparación",
    "Inventario",
    "IXC shipping_LengthEach",
    "IXC shipping_widthEach",
    "IXC shipping_heightEach",
    "IXC shipping_weightEach",
    "IXC shipping_volumeEach",
    "IXC otros metafields (JSON)",
]


def ixc_metafield_map(vnode: dict[str, Any]) -> dict[str, str]:
    out: dict[str, str] = {}
    block = vnode.get("metafields") or {}
    for edge in block.get("edges") or []:
        if not isinstance(edge, dict):
            continue
        node = edge.get("node") or {}
        k = node.get("key")
        if k:
            out[str(k)] = node.get("value") if node.get("value") is not None else ""
    return out


def ixc_metafield_columns(m: dict[str, str]) -> list[Any]:
    cells: list[Any] = []
    for key in IXC_SHIPPING_KEYS:
        cells.append(m.get(key, ""))
    rest = {k: v for k, v in m.items() if k not in IXC_SHIPPING_KEYS}
    cells.append(json.dumps(rest, ensure_ascii=False, separators=(",", ":")) if rest else "")
    return cells


def gid_numeric(gid: str | None) -> str:
    if not gid or not isinstance(gid, str):
        return ""
    m = re.search(r"/(\d+)\s*$", gid)
    return m.group(1) if m else gid


def variant_to_row(product: dict[str, Any], vnode: dict[str, Any]) -> list[Any]:
    opts = vnode.get("selectedOptions") or []
    ovals: list[str | None] = [None, None, None]
    for i, o in enumerate(opts[:3]):
        if isinstance(o, dict):
            ovals[i] = o.get("value")

    inv = vnode.get("inventoryQuantity")
    if inv is None:
        inv_str = ""
    else:
        inv_str = int(inv) if isinstance(inv, (int, float)) else str(inv)

    ixm = ixc_metafield_map(vnode)

    return [
        gid_numeric(product.get("id")),
        product.get("title") or "",
        product.get("handle") or "",
        str(product.get("status") or ""),
        product.get("vendor") or "",
        product.get("productType") or "",
        gid_numeric(vnode.get("id")),
        vnode.get("title") or "",
        vnode.get("sku") or "",
        vnode.get("barcode") or "",
        ovals[0] or "",
        ovals[1] or "",
        ovals[2] or "",
        vnode.get("price") or "",
        vnode.get("compareAtPrice") or "",
        inv_str,
        *ixc_metafield_columns(ixm),
    ]


def collect_variant_nodes(payload: dict[str, Any]) -> list[dict[str, Any]]:
    data = payload.get("data") or {}
    root = data.get("product")
    if not root:
        return []
    var_block = root.get("variants") or {}
    edges = var_block.get("edges") or []
    return [e.get("node") or {} for e in edges if isinstance(e, dict)]


def fetch_all_variant_nodes_for_product(
    conn: http.client.HTTPSConnection,
    api_version: str,
    token: str,
    product_id: str,
    first_block: dict[str, Any],
) -> list[dict[str, Any]]:
    """Primera página ya en first_block['variants']; pide el resto si hace falta."""
    variants_root = first_block.get("variants") or {}
    nodes: list[dict[str, Any]] = []
    edges = variants_root.get("edges") or []
    nodes.extend(e.get("node") or {} for e in edges if isinstance(e, dict))

    pinfo = variants_root.get("pageInfo") or {}
    vcursor = pinfo.get("endCursor")
    while pinfo.get("hasNextPage") and vcursor:
        time.sleep(0.15)
        resp = post_graphql(
            conn,
            api_version,
            token,
            VARIANTS_PAGE_QUERY,
            {"id": product_id, "cursor": vcursor},
        )
        if resp.get("errors"):
            raise RuntimeError(json.dumps(resp["errors"], indent=2))
        chunk = collect_variant_nodes(resp)
        if not chunk:
            break
        nodes.extend(chunk)
        data = resp.get("data") or {}
        prod = data.get("product") or {}
        vroot = prod.get("variants") or {}
        pinfo = vroot.get("pageInfo") or {}
        vcursor = pinfo.get("endCursor")

    return nodes


def iter_product_variant_rows(
    shop: str,
    token: str,
    api_version: str,
    page_delay: float,
) -> list[list[Any]]:
    host = normalize_shop(shop)
    rows: list[list[Any]] = []
    cursor: str | None = None
    has_next = True

    conn = http.client.HTTPSConnection(host, timeout=180)
    try:
        page = 0
        while has_next:
            page += 1
            if page > 1 and page_delay > 0:
                time.sleep(page_delay)

            payload = post_graphql(conn, api_version, token, PRODUCTS_PAGE_QUERY, {"cursor": cursor})
            if payload.get("errors"):
                raise RuntimeError(json.dumps(payload["errors"], indent=2))

            data = payload.get("data") or {}
            products = data.get("products") or {}
            edges = products.get("edges") or []

            for edge in edges:
                product = (edge.get("node") or {}) if isinstance(edge, dict) else {}
                pid = product.get("id")
                if not pid:
                    continue

                variant_nodes = fetch_all_variant_nodes_for_product(
                    conn, api_version, token, pid, product
                )
                if not variant_nodes:
                    rows.append(variant_to_row(product, {}))
                    continue
                for vnode in variant_nodes:
                    rows.append(variant_to_row(product, vnode))

            pinfo = products.get("pageInfo") or {}
            has_next = bool(pinfo.get("hasNextPage"))
            cursor = edges[-1].get("cursor") if edges else None
            if has_next and not cursor:
                raise RuntimeError("products: hasNextPage sin cursor")
            print(f"Página productos {page}, filas acumuladas: {len(rows)}", flush=True)
    finally:
        conn.close()

    return rows


def write_xlsx(path: Path, headers: list[str], rows: list[list[Any]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise SystemExit(
            "Falta openpyxl. Instale con: py -3 -m pip install openpyxl\n" + str(e)
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Variantes"
    ws.append(headers)
    for r in rows:
        ws.append(list(r))

    ws.freeze_panes = "A2"
    last_row = max(1, len(rows) + 1)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    widths = (
        14,
        36,
        22,
        12,
        18,
        22,
        14,
        28,
        18,
        18,
        14,
        14,
        14,
        12,
        18,
        10,
        22,
        22,
        22,
        22,
        22,
        40,
    )
    for i in range(1, len(headers) + 1):
        w = widths[i - 1] if i <= len(widths) else 18
        ws.column_dimensions[get_column_letter(i)].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    p = argparse.ArgumentParser(
        description="Exporta productos y variantes Shopify a Excel (.xlsx)."
    )
    p.add_argument("--shop", default=os.environ.get("SHOPIFY_SHOP", ""))
    p.add_argument("--token", default=os.environ.get("SHOPIFY_ACCESS_TOKEN", ""))
    p.add_argument("--api-version", default="2024-01")
    p.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path(__file__).resolve().parent / "shopify_catalogo_variantes.xlsx",
        help="Ruta del .xlsx de salida",
    )
    p.add_argument(
        "--page-delay",
        type=float,
        default=0.25,
        help="Pausa en segundos entre páginas de productos (rate limit)",
    )
    args = p.parse_args()

    if not args.shop or not args.token:
        print("Faltan --shop / --token o variables de entorno.", file=sys.stderr)
        return 1

    try:
        rows = iter_product_variant_rows(args.shop, args.token, args.api_version, args.page_delay)
    except (OSError, RuntimeError, json.JSONDecodeError) as e:
        print(e, file=sys.stderr)
        return 1

    write_xlsx(args.output, HEADERS_ES, rows)
    print(f"Listo: {len(rows)} filas de variantes -> {args.output.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
