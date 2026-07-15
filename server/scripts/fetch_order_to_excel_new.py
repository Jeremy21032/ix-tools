"""
fetch_order_to_excel_new.py
Consulta la API de consumerOrder para cada número de orden
y genera un Excel + JSON UI con: orderNumber, customerOrderNumber, netsuiteOrderNumber
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

import pandas as pd
import requests

# Env: CONSUMER_ORDER_BASE_URL
BASE_URL = os.environ.get(
    "CONSUMER_ORDER_BASE_URL",
    "https://ixc-ecom-fema-orchestrator-fn-prod.azurewebsites.net/api/consumerOrder",
)

FIELDS = ["orderNumber", "customerOrderNumber", "netsuiteOrderNumber"]
REPORT_COLUMNS = ["inputOrderId", *FIELDS, "error"]


def load_order_ids(orders: str | None, orders_file: Path | None) -> list[str]:
    ids: list[str] = []
    if orders:
        ids.extend(p.strip() for p in orders.split(",") if p.strip())
    if orders_file:
        if not orders_file.is_file():
            raise FileNotFoundError(f"No existe el archivo de órdenes: {orders_file}")
        raw = orders_file.read_text(encoding="utf-8")
        parts = re.split(r"[,\n\r]+", raw)
        ids.extend(p.strip() for p in parts if p.strip())
    return list(dict.fromkeys(ids))


def fetch_order(order_id: str) -> dict:
    url = f"{BASE_URL}/{order_id}"
    try:
        resp = requests.get(url, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        return {
            **{field: data.get(field, None) for field in FIELDS},
            "error": "",
        }
    except requests.exceptions.HTTPError as e:
        status = e.response.status_code if e.response is not None else "?"
        print(f"  HTTP error {order_id}: {e}")
        return {
            **{field: "" for field in FIELDS},
            "error": f"ERROR_HTTP_{status}",
        }
    except Exception as e:
        print(f"  Error {order_id}: {e}")
        return {**{field: "" for field in FIELDS}, "error": "ERROR"}


def export_json(rows: list[dict], output_path: Path) -> None:
    payload = {
        "columns": REPORT_COLUMNS,
        "rows": [{col: row.get(col, "") for col in REPORT_COLUMNS} for row in rows],
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Consulta consumerOrder por lista de IDs y exporta Excel."
    )
    parser.add_argument(
        "--orders",
        default="",
        help="IDs de orden separados por coma.",
    )
    parser.add_argument(
        "--orders-file",
        type=Path,
        help="Archivo con IDs (uno por línea o separados por coma).",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("orders_result.xlsx"),
        help="Ruta del Excel de salida.",
    )
    parser.add_argument(
        "--json-out",
        type=Path,
        default=None,
        help="Si se indica, también escribe JSON columns+rows para la UI.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        order_ids = load_order_ids(args.orders or None, args.orders_file)
    except FileNotFoundError as exc:
        print(exc, file=sys.stderr)
        return 1

    if not order_ids:
        print(
            "Indique --orders o --orders-file con al menos un ID de orden.",
            file=sys.stderr,
        )
        return 1

    total = len(order_ids)
    rows = []

    print(f"Consultando {total} órdenes...\n")
    for i, order_id in enumerate(order_ids, 1):
        print(f"[{i:>3}/{total}] {order_id}", end=" ... ")
        fetched = fetch_order(order_id)
        row = {
            "inputOrderId": order_id,
            "orderNumber": fetched.get("orderNumber") or "",
            "customerOrderNumber": fetched.get("customerOrderNumber") or "",
            "netsuiteOrderNumber": fetched.get("netsuiteOrderNumber") or "",
            "error": fetched.get("error") or "",
        }
        rows.append(row)
        print("OK" if not row["error"] else "FAIL")
        time.sleep(0.15)

    df = pd.DataFrame(rows, columns=REPORT_COLUMNS)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Órdenes")

        ws = writer.sheets["Órdenes"]

        from openpyxl.styles import Font, PatternFill, Alignment

        header_font = Font(name="Arial", bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", start_color="1F4E79")
        header_align = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        body_font = Font(name="Arial", size=10)
        for excel_row in ws.iter_rows(min_row=2):
            for cell in excel_row:
                cell.font = body_font

    print(f"\nExcel generado: {args.output}  ({total} filas)")

    json_out = args.json_out or args.output.with_suffix(".json")
    export_json(rows, json_out)
    print(f"JSON UI: {json_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
